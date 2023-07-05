import openpyxl
from openpyxl import load_workbook
import os

#initial setup
path = "C:/Users/haruk/Downloads/226347839_0_青少年校园霸凌与心理健康调查_208_207.xlsx"
output = os.path.abspath('.')
wb = load_workbook(path)
sheet = wb['Sheet1']
sheet2 = wb.create_sheet("Data")
print(wb.sheetnames)
m_row = sheet.max_row

#calculate IAS sheet score
s1_start = (sheet['J2'].column)
s1_end = (sheet['X2'].column)
i = s1_start
j = 2
sta = 0

while(j <= m_row):
    while(i <= s1_end):
        score = sheet.cell(row=j,column=i).value
        tmp = i-s1_start+1

        if tmp in (3,6,10,15):  #reverse scores caculate
            if(score == "极其相符"):
                score = 1
            elif(score == "非常相符"):
                score = 2
            elif(score == "中等相符"):
                score = 3
            elif(score == "部分相符"):
                score = 4
            elif(score == "完全不符"):
                score = 5
        
        else:    #regular calculate
            if(score == "极其相符"):
                score = 5
            elif(score == "非常相符"):
                score = 4
            elif(score == "中等相符"):
                score = 3
            elif(score == "部分相符"):
                score = 2
            elif(score == "完全不符"):
                score = 1
        
        sta = sta + score
        #write scores to new sheet
        sheet2.cell(row=j,column=i-8,value=score).value
        sheet2.cell(row=1,column=i-8,value=tmp).value
        i = i+1

    sheet2.cell(row=j,column=s1_end - 7,value=sta).value
    i = s1_start
    sta = 0
    j = j+1

#calculate SAS sheet score
s2_start = (sheet['Y2'].column)
s2_end = (sheet['AR2'].column)
i = s2_start
j = 2
sta = 0

while(j <= m_row):
    while(i <= s2_end):
        score = sheet.cell(row=j,column=i).value
        tmp = i-s2_start+1

        if tmp in (5,9,13,17,19):  #reverse scores caculate
            if(score == "绝大部分或全部时间"):
                score = 1
            elif(score == "相当多时间"):
                score = 2
            elif(score == "小部分时间"):
                score = 3
            elif(score == "没有或很少时间"):
                score = 4
        
        else:    #regular calculate
            if(score == "绝大部分或全部时间"):
                score = 4
            elif(score == "相当多时间"):
                score = 3
            elif(score == "小部分时间"):
                score = 2
            elif(score == "没有或很少时间"):
                score = 1
        
        sta = sta + score
        #write scores to new sheet
        sheet2.cell(row=j,column=i-7,value=score).value
        sheet2.cell(row=1,column=i-7,value=tmp).value
        i = i+1

    sheet2.cell(row=j,column=s2_end - 6,value=sta).value
    i = s2_start
    sta = 0
    j = j+1

#calculate SDS sheet score
s3_start = (sheet['AS2'].column)
s3_end = (sheet['BL2'].column)
i = s3_start
j = 2
sta = 0

while(j <= m_row):
    while(i <= s3_end):
        score = sheet.cell(row=j,column=i).value
        tmp = i-s3_start+1

        if tmp in (2,5,6,11,12,14,16,17,18,20):  #reverse scores caculate
            if(score == "绝大部分或全部时间"):
                score = 1
            elif(score == "相当多时间"):
                score = 2
            elif(score == "小部分时间"):
                score = 3
            elif(score == "没有或很少时间"):
                score = 4
        
        else:    #regular calculate
            if(score == "绝大部分或全部时间"):
                score = 4
            elif(score == "相当多时间"):
                score = 3
            elif(score == "小部分时间"):
                score = 2
            elif(score == "没有或很少时间"):
                score = 1
        
        sta = sta + score
        #write scores to new sheet
        sheet2.cell(row=j,column=i-6,value=score).value
        sheet2.cell(row=1,column=i-6,value=tmp).value
        i = i+1

    sheet2.cell(row=j,column=s3_end - 5,value=sta).value
    i = s3_start
    sta = 0
    j = j+1

#calculate total
i = 2
while(i <= m_row):
    score1 = sheet2.cell(row=i,column=s1_end - 7).value
    score2 = sheet2.cell(row=i,column=s2_end - 6).value
    score3 = sheet2.cell(row=i,column=s3_end - 5).value
    total = score1+score2+score3
    sheet2.cell(row=i,column=s3_end -4,value=total).value
    sheet2.cell(row=i,column=1,value=i).value
    i = i+1

#write comments
sheet2.cell(row=1,column=1,value="No./Answer No.").value
sheet2['Q1'] = "IAS Total"
sheet2['AL1'] = "SAS Total"
sheet2['BG1'] = "SDS Total"
sheet2['BH1'] = "Total"

wb.save("data.xlsx")
wb.close()